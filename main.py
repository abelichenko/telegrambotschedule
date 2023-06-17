# -*- coding: cp1251 -*-
import telebot
import ast
import time
from telebot import types
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
import wget


bot = telebot.TeleBot("5886031620:AAGouOzMeOtyaYbAuhMnerLD8vqhpodU2dY")
current_datetime = str(datetime.now().date())
adminUser = {"abelichenko", "test"}


adminMode = True
inputOn = 0
lessonName = ""
audience = ""
teacher = ""
editGroup = ""
editDays = ""
editNumRow = ""
wb = openpyxl.reader.excel.load_workbook(filename="table_base.xlsx")

wb.active = 0
sheet = wb.active
lessonNumber = {"1_pair": "Перша пара", "2_pair": "Друга пара", "3_pair": "Третя пара", "4_pair": "Четверта пара", "5_pair": "П'ята пара", "6_pair": "Шоста пара", "20": "На головну"}
daysNumber = {"1_1": "Понеділок", "2_2": "Вівторок", "3_3": "Середа", "4_4": "Четвер", "5_5": "Пятниця", "20": "На головну"}
allGroupList = {"1": "УПІ9 -22-10", "2": "ФБС9/11-21/22-20", "3": "ОО9/11-21/22-21", "4": "ПТБ9/11-21/22-22", "5": "М9-21-25", "6": "ІПЗ9/11-21/22-26", "7": "ПВ9/11-21/22-28", "8": "ФБС9/11-20/21-30", "9": "ПВ9/11-21/22-28", "10": "ФБС9/11-20/21-30", "11": "ОО9/11-20/21-31", "12": "ПТБ9/11-20/21-32", "13": "М9/11-20/21-35", "14": "ІПЗ9/11-20/21-36", "15": "ПВ9/11-20/21-38", "16": "ІПЗ9-19-46/ІПЗ11-20-47", "17": "ПВ9-19-48/ПВ11-20-49", "18": "Завантажити таблицю", "19": "Відновити таблицю"}

buttonIconEdit = u"\u270E"
buttonReadUser = u"\u2709"

def backup():
    os.remove('table_base.xlsx')
    url = f'https://drive.google.com/uc?export=download&id=1c23XbgbEGESwD_3uIYYIiFpen1BFe5PY'
    wget.download(url)
    print('Good')
    return


def makeKeyboard(acess):
    
    
    if(acess):        
        markup = types.InlineKeyboardMarkup()
    
        for key, value in allGroupList.items():
            if(key == "19" or key == "18"):
                markup.add(types.InlineKeyboardButton(text=value,
                                                      callback_data="['value', '" + value + "', '" + key + "']"),
                types.InlineKeyboardButton(text=buttonReadUser,
                                           callback_data="['key', '" + key + "']"))
            else:
                markup.add(types.InlineKeyboardButton(text=value,
                                                      callback_data="['value', '" + value + "', '" + key + "']"),
                           types.InlineKeyboardButton(text=buttonIconEdit,
                                                      callback_data="['key', '" + key + "']"))
        return markup
    else:
        markup = types.InlineKeyboardMarkup()

        for key, value in allGroupList.items():
            if(key != "19" and key != "20"):
                markup.add(types.InlineKeyboardButton(text=value,
                                                      callback_data="['value', '" + value + "', '" + key + "']"),
                           types.InlineKeyboardButton(text=buttonReadUser,
                                                      callback_data="['key', '" + key + "']"))
        return markup

def makeKeyboard_step_2():
    markup = types.InlineKeyboardMarkup()

    for key, value in daysNumber.items():
        markup.add(types.InlineKeyboardButton(text=value,
                                              callback_data="['value', '" + value + "', '" + key + "']"),
        types.InlineKeyboardButton(text=buttonIconEdit,
                                   callback_data="['key', '" + key + "']"))


    return markup

def makeKeyboard_step_3():
    markup = types.InlineKeyboardMarkup()

    for key, value in lessonNumber.items():
        markup.add(types.InlineKeyboardButton(text=value,
                                              callback_data="['value', '" + value + "', '" + key + "']"),
        types.InlineKeyboardButton(text=buttonIconEdit,
                                   callback_data="['key', '" + key + "']"))


    return markup

@bot.message_handler(commands=['adminmode'])
def admMode(message):
    global adminMode
    acess = False
    for char in adminUser:
        if (message.from_user.username == char):
            acess = True
    if(acess != True):
        bot.send_message(chat_id=message.chat.id, text="У вас не достатньо прав на виконання команди!")
        return
    if(adminMode and acess):
        adminMode = False
        bot.send_message(chat_id=message.chat.id, text="Режим адміністрування деактивовано!")
    elif(adminMode != True and acess):
        adminMode = True
        bot.send_message(chat_id=message.chat.id, text="Режим адміністрування активовано!")


@bot.message_handler(commands=['start'])
def handle_command_adminwindow(message):
    acess = False
    for char in adminUser:
        if (message.from_user.username == char):
            acess = True
    if(acess == True and adminMode == False):
        acess = False
            
    bot.send_message(chat_id=message.chat.id,
                     text="Список групп колледжу:",
                     reply_markup=makeKeyboard(acess),
                     parse_mode='HTML')

@bot.callback_query_handler(func=lambda call: True)
def handle_query(call):
    if (call.data.startswith("['value'")):
        print(f"call.data : {call.data} , type : {type(call.data)}")
        print(f"ast.literal_eval(call.data) : {ast.literal_eval(call.data)} , type : {type(ast.literal_eval(call.data))}")
        valueFromCallBack = ast.literal_eval(call.data)[1]
        keyFromCallBack = ast.literal_eval(call.data)[2]
        bot.answer_callback_query(callback_query_id=call.id,
                              show_alert=True,
                              text="Вы выбрали " + valueFromCallBack + " и ключь его " + keyFromCallBack)
    print(ast.literal_eval(call.data)[1])
    if (call.data.startswith("['key'")):
        keyFromCallBack = ast.literal_eval(call.data)[1]
        valueFromCallBack = ast.literal_eval(call.data)[1]
        i = 0
        acess = False
        for char in adminUser:
            if (call.from_user.username == char):
                acess = True
                #bot.send_message(call.message.chat.id, acess)
        if acess:
            if (keyFromCallBack != "18" and keyFromCallBack != "19" and keyFromCallBack != "20"):
                for key, value in daysNumber.items():
                    if(keyFromCallBack == key):
                        i=+1
                if(i >= 1):
                    editTable(call, keyFromCallBack)
                else:
                    for key, value in lessonNumber.items():
                        if (keyFromCallBack == key):
                            i = +1
                    if (i >= 1):
                        numPair(call, keyFromCallBack)
                    else:
                        viewTable(call, keyFromCallBack)
            elif (keyFromCallBack == "18"):
                bot.send_message(call.message.chat.id, "Остання створенна таблиця:")
                f = open("table_base.xlsx", "rb")
                bot.send_document(call.message.chat.id, f)
                return
            elif (keyFromCallBack == "19"):
                backup()
            elif(keyFromCallBack == "20"):
                acess = False
                for char in adminUser:
                    if (call.from_user.username == char):
                        acess = True
                bot.edit_message_text(chat_id=call.message.chat.id,
                                      text="Список групп колледжу: ",
                                      message_id=call.message.message_id,
                                      reply_markup=makeKeyboard(acess),
                                      parse_mode='HTML')
        else:
            if(keyFromCallBack != "18" and keyFromCallBack != "19" and keyFromCallBack != "20"):
                makeTableForGroup(call, keyFromCallBack)
            else:
                bot.send_message(call.message.chat.id, "Повний розклад:")
                f = open("table_base.xlsx", "rb")
                bot.send_document(call.message.chat.id, f)


def makeTableForGroup(call, key):
    global editGroup
    search_text = ""
    for keys, value in allGroupList.items():
        if keys == key:
            search_text = value
    if(key == "20"):
        acess = False
        for char in adminUser:
            if (call.from_user.username == char):
                acess = True
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text="Список групп колледжу: ",
                              message_id=call.message.message_id,
                              reply_markup=makeKeyboard(acess),
                              parse_mode='HTML')
    row_max = sheet.max_row
    column_max = sheet.max_column
    row_min = 1  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца

    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)

            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min

            data_from_cell = sheet[word_cell].value
            data_from_cell = str(data_from_cell)
            # print(data_from_cell)
            regular = search_text
            result = re.findall(regular, data_from_cell)
            if len(result) > 0:
                print('Нашли в ячейке:', word_cell)
                editGroup = word_column
            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1

    #sheet[editNumRow].value = str(lessonName + "\n" + teacher + "\n" + "ауд." + audience)



    wbCopy = openpyxl.reader.excel.load_workbook(filename="form.xlsx")

    wbCopy.active = 0
    sheetCopy = wbCopy.active
    sheetCopy['C1'].value = str(search_text)
    i = 0
    nums_original = 9
    nums_copy = 2
    editGroupRaw = editGroup
    editCopyRaw = "C"
    editGroup = editGroupRaw + str(nums_original)
    editCopy = "C" + str(nums_copy)
    while i < 26:
        editString = sheet[editGroup].value
        j = 0
        editaccept = 0
        ready = True
        if (editString is None):
            ready = False
        if(ready):
            editStringNextStep = list(editString)
            while j < len(editString):
                if j+3 < len(editString):
                    if(editStringNextStep[j] == " " and editStringNextStep[j + 1] == " "):
                        editStringNextStep[j] = "j"
                        print(editStringNextStep)
                        editaccept = editaccept + 1
                j = j + 1
            if(editaccept > 0):
                editString = ""
                for s in editStringNextStep:
                    editString += str(s)

                editString = editString.replace("j", "")
            print(editString)
            sheetCopy[editCopy].value = editString
        else:
            sheetCopy[editCopy].value = editString
        nums_original = nums_original + 2
        nums_copy = nums_copy + 2
        editGroup = editGroupRaw + str(nums_original)
        editCopy = editCopyRaw + str(nums_copy)
        #print(editCopy, editGroup)
        i = i + 1
    print("Save!")

    wbCopy.save("form.xlsx")

    bot.send_message(call.message.chat.id, "Ваш розклад:")
    f = open("form.xlsx", "rb")
    bot.send_document(call.message.chat.id, f)
    return


def viewTable(call, key):
    global editGroup

    if(key == "20"):
        acess = False
        for char in adminUser:
            if (call.from_user.username == char):
                acess = True
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text="Список групп колледжу: ",
                              message_id=call.message.message_id,
                              reply_markup=makeKeyboard(acess),
                              parse_mode='HTML')

    search_text = ""
    for keys, value in allGroupList.items():
        if (key == keys):
            i = str(value)
            search_text = i
    row_max = sheet.max_row
    column_max = sheet.max_column
    row_min = 1  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца

    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)

            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min

            data_from_cell = sheet[word_cell].value
            data_from_cell = str(data_from_cell)
            # print(data_from_cell)
            regular = search_text
            result = re.findall(regular, data_from_cell)
            if len(result) > 0:
                print('Нашли в ячейке:', word_cell)
                editGroup = word_column
            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1
    bot.edit_message_text(chat_id=call.message.chat.id,
                    text="Оберіть день: ",
                    message_id=call.message.message_id,
                    reply_markup=makeKeyboard_step_2(),
                    parse_mode='HTML')

def editTable(call, key):
    global editDays

    if(key == "20"):
        acess = False
        for char in adminUser:
            if (call.from_user.username == char):
                acess = True
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text="Список групп колледжу: ",
                              message_id=call.message.message_id,
                              reply_markup=makeKeyboard(acess),
                              parse_mode='HTML')

    if(key == "1_1"):
        editDays = '9'
    elif(key == "2_2"):
        editDays = '21'
    elif (key == "3_3"):
        editDays = '31'
    elif (key == "4_4"):
        editDays = '41'
    elif (key == "5_5"):
        editDays = '51'
    bot.edit_message_text(chat_id=call.message.chat.id,
                    text="Оберіть день: ",
                    message_id=call.message.message_id,
                    reply_markup=makeKeyboard_step_3(),
                    parse_mode='HTML')


def numPair(call, key):
    global editDays
    global inputOn
    inputOn = 3

    if(key == "20"):
        acess = False
        for char in adminUser:
            if (call.from_user.username == char):
                acess = True
        bot.edit_message_text(chat_id=call.message.chat.id,
                              text="Список групп колледжу: ",
                              message_id=call.message.message_id,
                              reply_markup=makeKeyboard(acess),
                              parse_mode='HTML')

    if (key == "1_pair"):
        editDays = str(int(editDays) + 0)
    elif (key == "2_pair"):
        editDays = str(int(editDays) + 2)
    elif (key == "3_pair"):
        editDays = str(int(editDays) + 4)
    elif (key == "4_pair"):
        editDays = str(int(editDays) + 6)
    elif (key == "5_pair"):
        editDays = str(int(editDays) + 8)
    elif (editDays <= 9 and key == "6_pair"):
        editDays = str(int(editDays) + 10)

    bot.send_message(call.message.chat.id, "Введіть назву предмета")


@bot.message_handler(content_types=['text'])
def input(message):
    global inputOn
    global audience
    global teacher
    global lessonName
    global editDays
    global editGroup
    global editNumRow
    print(inputOn)
    editNumRow = editGroup + editDays
    if(inputOn == 3):
        print("Step 2")
        lessonName = message.text
        bot.send_message(message.chat.id, "Введіть ініціали викладача")

    if(inputOn == 2):
        print("Step 3")
        teacher = message.text
        bot.send_message(message.chat.id, "Введіть номер аудиторії")
    if(inputOn == 1):
        print("Step 4")
        audience = message.text
        sheet[editNumRow].value = str(lessonName + "\n" + teacher + "\n" + "ауд." + audience)
        wb.save("table_base.xlsx")
        bot.send_message(message.chat.id, "Таблицю успішно створено")
        inputOn = 0
        handle_command_adminwindow(message)

    inputOn = inputOn - 1
while True:
    try:
        bot.polling(none_stop=True, interval=0, timeout=0)
    except:
        time.sleep(10)
